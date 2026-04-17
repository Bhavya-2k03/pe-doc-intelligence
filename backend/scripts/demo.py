"""
End-to-end demo: feed clauses -> LLM interprets -> timelines built -> visualized.

Run:
    python demo.py

What this does:
  1. Defines a small set of SAMPLE CLAUSES (you can edit these freely)
  2. Sends each clause to the LLM (clause_interpreter)
  3. Shows you exactly what the LLM returned (ClauseInstruction objects)
  4. Executes them against seed timelines in order
  5. Prints a readable timeline for every field
  6. Exports everything to demo_output/results.xlsx

You can edit SAMPLE_CLAUSES below to test any text you want.
"""
from __future__ import annotations

import asyncio
import json
import os
import sys
from datetime import date
from pathlib import Path
from typing import Any

# Ensure backend/ (the parent of scripts/) is on sys.path so that
# `from engine.xxx import ...` resolves when this script is executed
# directly as `python scripts/demo.py`.
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openai import AsyncOpenAI

load_dotenv()

# ─────────────────────────────────────────────────────────────────────────────
# EDIT THIS SECTION — add/remove/change any clauses you want to test
# ─────────────────────────────────────────────────────────────────────────────

SAMPLE_CLAUSES = [
    {
        "clause_text": (
            "The management fee rate shall be reduced from 2.0% to 1.75% per annum with effect from January 1, 2026."
        ),
        "source_signed_date": "2025-12-01",
        "source_effective_date": "2026-01-01",
        "doc_type": "side_letter",
        "email_source_id": "demo-e001",
    },
    {
        "clause_text": (
            "The management fee basis shall switch from committed capital to invested capital upon the end of the Investment Period."
        ),
        "source_signed_date": "2025-12-01",
        "source_effective_date": "2029-01-15",
        "doc_type": "side_letter",
        "email_source_id": "demo-e001",
    },
    {
        "clause_text": (
            "Notwithstanding the foregoing, the management fee shall not exceed 1.5% per annum at any time after January 1, 2027."
        ),
        "source_signed_date": "2025-12-15",
        "source_effective_date": "2027-01-01",
        "doc_type": "side_letter",
        "email_source_id": "demo-e002",
    },
    {
        "clause_text": (
            "The fee reduction described in Section 2 shall be deferred until the second anniversary of the Fund's Final Closing Date."
        ),
        "source_signed_date": "2025-12-20",
        "source_effective_date": None,
        "doc_type": "side_letter",
        "email_source_id": "demo-e003",
    },
    {
        "clause_text": (
            "This letter is for informational purposes only and does not modify any economic terms of the Limited Partnership Agreement."
        ),
        "source_signed_date": "2025-11-01",
        "source_effective_date": None,
        "doc_type": "side_letter",
        "email_source_id": "demo-e004",
    },
]

# Evaluation date — the date we're querying timelines for
EVALUATION_DATE = date(2030, 6, 1)

# LPA seed timelines (the "before any side letters" baseline)
SEED_TIMELINES = {
    "management_fee_rate": [
        {"date": "2024-01-15", "value": 2.0, "source": "LPA"},
    ],
    "management_fee_basis": [
        {"date": "2024-01-15", "end_date": "2029-01-15",
         "value": "committed_capital", "source": "LPA"},
        {"date": "2029-01-15", "value": "invested_capital", "source": "LPA"},
    ],
    "carried_interest_rate": [
        {"date": "2024-01-15", "value": 20, "source": "LPA"},
    ],
    "preferred_return_rate": [
        {"date": "2024-01-15", "value": 8, "source": "LPA"},
    ],
    "fund_final_closing_date": [
        {"date": "2024-01-15", "value": "2024-12-15", "source": "LPA"},
    ],
    "fund_investment_end_date": [
        {"date": "2024-01-15", "value": "2029-01-15", "source": "LPA"},
    ],
}

OUTPUT_DIR = "demo_output"

# ─────────────────────────────────────────────────────────────────────────────
# Logging: print to terminal AND write to text file simultaneously
# ─────────────────────────────────────────────────────────────────────────────

import builtins

_log_file = None  # opened in main()

def _tee_print(*args, **kwargs):
    """Drop-in replacement for print() that mirrors output to the log file."""
    builtins.__original_print__(*args, **kwargs)
    if _log_file is not None:
        kwargs.pop("file", None)
        builtins.__original_print__(*args, file=_log_file, **kwargs)

# Monkey-patch print so all calls (including helpers) go through _tee_print
builtins.__original_print__ = builtins.print
builtins.print = _tee_print


# ─────────────────────────────────────────────────────────────────────────────
# Engine imports
# ─────────────────────────────────────────────────────────────────────────────

from engine.clause_interpreter import interpret_clause
from engine.models import ClauseInstruction
from engine.timeline_engine import (
    EvaluationContext,
    FieldTimeline,
    TimelineEntry,
    execute,
)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _sep(char="─", width=80):
    print(char * width)


def _print_instruction(i: int, instruction: ClauseInstruction):
    """Pretty-print one ClauseInstruction."""
    print(f"  [{i}] action      : {instruction.action}")
    print(f"       affected    : {instruction.affected_field or '(none)'}")
    if instruction.action == "SET":
        print(f"       value_expr  : {instruction.value_expr}")
        print(f"       eff_date    : {instruction.effective_date_expr}")
        print(f"       eff_end     : {instruction.effective_end_date_expr}")
        print(f"       condition   : {instruction.condition_ast}")
    elif instruction.action == "ADJUST":
        print(f"       delta       : {instruction.value_expr}")
        print(f"       direction   : {instruction.adjust_direction}")
        print(f"       eff_date    : {instruction.effective_date_expr}")
    elif instruction.action == "CONSTRAIN":
        print(f"       bound       : {instruction.value_expr}")
        print(f"       type        : {instruction.constraint_type}")
        print(f"       active_from : {instruction.effective_date_expr}")
        print(f"       active_until: {instruction.effective_end_date_expr}")
    elif instruction.action == "GATE":
        print(f"       gate_target : {instruction.gate_target}")
        print(f"       scope_mode  : {instruction.gate_scope_mode}")
        print(f"       move_to     : {instruction.gate_move_to_date_expr}")
        print(f"       condition   : {instruction.condition_ast}")
    elif instruction.action == "NO_ACTION":
        print(f"       reason      : {instruction.no_action_reason}")
    elif instruction.action == "MANUAL_REVIEW":
        print(f"       reason      : {instruction.manual_review_reason}")


def _build_seed_timelines() -> dict[str, FieldTimeline]:
    timelines: dict[str, FieldTimeline] = {}
    for field_name, entries in SEED_TIMELINES.items():
        ft = FieldTimeline()
        for e in entries:
            sd = e["date"]
            if isinstance(sd, str):
                sd = date.fromisoformat(sd)
            ed = e.get("end_date")
            if isinstance(ed, str):
                ed = date.fromisoformat(ed)
            ft.insert_entry(TimelineEntry(
                date=sd, end_date=ed, value=e["value"],
                source_clause_text=e.get("source", "LPA"),
                entry_type="SET",
            ))
        timelines[field_name] = ft
    return timelines


def _print_timelines(timelines: dict[str, FieldTimeline]):
    for field_name in sorted(timelines.keys()):
        ft = timelines[field_name]
        if not ft.entries:
            continue
        print(f"\n  {field_name}:")
        for e in ft.entries:
            end_str = str(e.end_date) if e.end_date else "open-ended"
            print(f"    {e.date} -> {end_str}  |  value={e.value}  |  "
                  f"src={e.source_clause_text[:55]}")
        # Also show what value_at(EVALUATION_DATE) returns
        val = ft.value_at(EVALUATION_DATE)
        print(f"    >> value on {EVALUATION_DATE}: {val}")

        if ft.constraints:
            print(f"    CONSTRAINTS:")
            for c in ft.constraints:
                af = f"from {c.active_from}" if c.active_from else ""
                print(f"      {c.type} {c.bound} {af}")


def _export_excel(
    clauses: list[dict],
    interpretation_results: list[tuple[dict, list[ClauseInstruction] | None]],
    timelines: dict[str, FieldTimeline],
    evaluation_date: date,
) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = Workbook()

    # ── Sheet 1: Clauses ─────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Input Clauses"
    h1_headers = ["#", "Clause Text", "Doc Type", "Signed Date",
                  "Effective Date", "Email Source"]
    for col, h in enumerate(h1_headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5496", end_color="2F5496",
                                fill_type="solid")
    for i, c in enumerate(clauses, 1):
        ws1.cell(row=i + 1, column=1, value=i)
        ws1.cell(row=i + 1, column=2, value=c["clause_text"])
        ws1.cell(row=i + 1, column=3, value=c.get("doc_type", ""))
        ws1.cell(row=i + 1, column=4, value=c.get("source_signed_date", ""))
        ws1.cell(row=i + 1, column=5, value=c.get("source_effective_date", "") or "")
        ws1.cell(row=i + 1, column=6, value=c.get("email_source_id", ""))
    ws1.column_dimensions["B"].width = 70
    for col in ["A", "C", "D", "E", "F"]:
        ws1.column_dimensions[col].width = 18

    # ── Sheet 2: LLM Interpretations ─────────────────────────────────
    ws2 = wb.create_sheet("LLM Interpretations")
    h2_headers = ["#", "Clause Text (truncated)", "Action", "Affected Field",
                  "Value / Delta", "Direction", "Constraint Type",
                  "Effective Date Expr", "Condition", "No-Action Reason",
                  "Manual Review Reason"]
    for col, h in enumerate(h2_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="375623", end_color="375623",
                                fill_type="solid")

    row = 2
    for clause, instructions in interpretation_results:
        if instructions is None:
            ws2.cell(row=row, column=1, value="?")
            ws2.cell(row=row, column=2, value=clause["clause_text"][:60])
            ws2.cell(row=row, column=3, value="FAILED")
            row += 1
            continue
        for instr in instructions:
            ws2.cell(row=row, column=1, value=clause.get("email_source_id",""))
            ws2.cell(row=row, column=2, value=clause["clause_text"][:80])
            ws2.cell(row=row, column=3, value=instr.action)
            ws2.cell(row=row, column=4, value=instr.affected_field or "")
            ws2.cell(row=row, column=5,
                     value=str(instr.value_expr) if instr.value_expr else "")
            ws2.cell(row=row, column=6, value=instr.adjust_direction or "")
            ws2.cell(row=row, column=7, value=instr.constraint_type or "")
            ws2.cell(row=row, column=8,
                     value=str(instr.effective_date_expr) if instr.effective_date_expr else "")
            ws2.cell(row=row, column=9,
                     value=str(instr.condition_ast) if instr.condition_ast else "")
            ws2.cell(row=row, column=10, value=instr.no_action_reason or "")
            ws2.cell(row=row, column=11, value=instr.manual_review_reason or "")

            # Color rows by action
            action_colors = {
                "SET": "D9EAD3",
                "ADJUST": "FCE5CD",
                "CONSTRAIN": "CFE2F3",
                "GATE": "EAD1DC",
                "NO_ACTION": "F3F3F3",
                "MANUAL_REVIEW": "FFF2CC",
            }
            fill_color = action_colors.get(instr.action, "FFFFFF")
            for col in range(1, len(h2_headers) + 1):
                ws2.cell(row=row, column=col).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color,
                    fill_type="solid",
                )
            row += 1
    ws2.column_dimensions["B"].width = 55
    for col_letter in ["C", "D", "F", "G"]:
        ws2.column_dimensions[col_letter].width = 18
    for col_letter in ["E", "H", "I", "J", "K"]:
        ws2.column_dimensions[col_letter].width = 30

    # ── Sheet 3: Final Timelines ──────────────────────────────────────
    ws3 = wb.create_sheet("Final Timelines")
    h3_headers = ["Field", "Start Date", "End Date", "Value",
                  f"Value on {evaluation_date}", "Source", "Entry Type"]
    for col, h in enumerate(h3_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="7B3F00", end_color="7B3F00",
                                fill_type="solid")

    row = 2
    for field_name in sorted(timelines.keys()):
        ft = timelines[field_name]
        if not ft.entries:
            continue
        current_val = ft.value_at(evaluation_date)
        first = True
        for e in ft.entries:
            ws3.cell(row=row, column=1, value=field_name if first else "")
            ws3.cell(row=row, column=2, value=str(e.date))
            ws3.cell(row=row, column=3,
                     value=str(e.end_date) if e.end_date else "open-ended")
            ws3.cell(row=row, column=4, value=str(e.value))
            ws3.cell(row=row, column=5,
                     value=str(current_val) if first else "")
            ws3.cell(row=row, column=6, value=e.source_clause_text[:60])
            ws3.cell(row=row, column=7, value=e.entry_type)
            if first:
                ws3.cell(row=row, column=1).font = Font(bold=True)
            if row % 2 == 0:
                for col in range(1, len(h3_headers) + 1):
                    ws3.cell(row=row, column=col).fill = PatternFill(
                        start_color="EAF2FB", end_color="EAF2FB", fill_type="solid",
                    )
            first = False
            row += 1

    for col_letter, width in zip(
        ["A", "B", "C", "D", "E", "F", "G"],
        [28, 14, 14, 18, 18, 55, 12],
    ):
        ws3.column_dimensions[col_letter].width = width

    # ── Sheet 4: Field Snapshot (value on evaluation_date) ───────────
    ws4 = wb.create_sheet(f"Snapshot {evaluation_date}")
    h4_headers = ["Field", f"Value on {evaluation_date}", "# Timeline Entries",
                  "Constraints"]
    for col, h in enumerate(h4_headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4A235A", end_color="4A235A",
                                fill_type="solid")
    for row, field_name in enumerate(sorted(timelines.keys()), 2):
        ft = timelines[field_name]
        val = ft.value_at(evaluation_date)
        constraints_str = "; ".join(
            f"{c.type} {c.bound}" for c in ft.constraints
        )
        ws4.cell(row=row, column=1, value=field_name)
        ws4.cell(row=row, column=2, value=str(val) if val is not None else "N/A")
        ws4.cell(row=row, column=3, value=len(ft.entries))
        ws4.cell(row=row, column=4, value=constraints_str or "none")
    for col_letter, width in zip(["A", "B", "C", "D"], [28, 20, 10, 40]):
        ws4.column_dimensions[col_letter].width = width

    path = os.path.join(OUTPUT_DIR, "results.xlsx")
    wb.save(path)
    return path


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

async def main():
    global _log_file
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    log_path = os.path.join(OUTPUT_DIR, "demo_run.txt")
    _log_file = open(log_path, "w", encoding="utf-8")

    try:
        await _run()
    finally:
        _log_file.close()
        _log_file = None
        print(f"  Log saved: {log_path}")


async def _run():
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        print("ERROR: OPENAI_API_KEY not found in .env")
        return

    client = AsyncOpenAI(api_key=api_key)

    print()
    _sep("=")
    print("  PE DOCUMENT INTELLIGENCE — DEMO")
    print(f"  Evaluation Date: {EVALUATION_DATE}")
    print(f"  Clauses to test: {len(SAMPLE_CLAUSES)}")
    _sep("=")

    # ── Step 1: Show input clauses ────────────────────────────────────
    print("\nINPUT CLAUSES:")
    _sep()
    for i, c in enumerate(SAMPLE_CLAUSES, 1):
        print(f"\n[{i}] {c['doc_type']} | signed {c['source_signed_date']} "
              f"| effective {c.get('source_effective_date') or '(none)'}")
        print(f"    \"{c['clause_text']}\"")

    # ── Step 2: Interpret each clause with the LLM ───────────────────
    print("\n\nLLM INTERPRETATIONS (calling GPT):")
    _sep()

    interpretation_results: list[tuple[dict, list[ClauseInstruction] | None]] = []

    for i, clause in enumerate(SAMPLE_CLAUSES, 1):
        print(f"\n[{i}] \"{clause['clause_text'][:70]}...\"")
        try:
            instructions = await interpret_clause(clause["clause_text"], client)
            interpretation_results.append((clause, instructions))
            print(f"  -> {len(instructions)} instruction(s) returned:")
            for j, instr in enumerate(instructions):
                _print_instruction(j, instr)
        except Exception as e:
            print(f"  -> FAILED: {e}")
            interpretation_results.append((clause, None))

    # ── Step 3: Build seed timelines ─────────────────────────────────
    print("\n\nSEED TIMELINES (LPA baseline, before any clauses):")
    _sep()
    timelines = _build_seed_timelines()
    _print_timelines(timelines)

    # ── Step 4: Execute instructions sequentially ────────────────────
    print("\n\nEXECUTING CLAUSE INSTRUCTIONS:")
    _sep()

    eval_ctx = EvaluationContext(
        evaluation_date=EVALUATION_DATE,
        document_date=None,
        fund_data={},
    )

    for clause, instructions in interpretation_results:
        if instructions is None:
            continue
        for instr in instructions:
            if instr.action in ("NO_ACTION", "MANUAL_REVIEW"):
                print(f"  SKIP [{instr.action}]: {clause['clause_text'][:60]}")
                continue

            # Set document_date for this clause
            eff_date_str = clause.get("source_effective_date") or \
                           clause.get("source_signed_date")
            if eff_date_str:
                eval_ctx.document_date = date.fromisoformat(eff_date_str)

            print(f"\n  EXECUTE {instr.action} on '{instr.affected_field}' "
                  f"| doc_date={eval_ctx.document_date}")
            execute(instr, timelines, eval_ctx)

    # ── Step 5: Show final timelines ──────────────────────────────────
    print("\n\nFINAL TIMELINES (after all clauses executed):")
    _sep()
    _print_timelines(timelines)

    # ── Step 6: Snapshot on evaluation_date ──────────────────────────
    print(f"\n\nFIELD VALUES ON {EVALUATION_DATE} (the evaluation date):")
    _sep()
    for field_name in sorted(timelines.keys()):
        val = timelines[field_name].value_at(EVALUATION_DATE)
        print(f"  {field_name:<35} = {val}")

    # ── Step 7: Export to Excel ───────────────────────────────────────
    print(f"\n\nEXPORTING TO EXCEL...")
    path = _export_excel(
        SAMPLE_CLAUSES, interpretation_results, timelines, EVALUATION_DATE
    )
    print(f"  Saved: {path}")
    print(f"\n  The Excel file has 4 sheets:")
    print(f"    1. Input Clauses    — your input text")
    print(f"    2. LLM Interpretations — what GPT returned (color-coded by action)")
    print(f"    3. Final Timelines  — all timeline entries for every field")
    print(f"    4. Snapshot {EVALUATION_DATE} — one value per field on evaluation date")
    print()


if __name__ == "__main__":
    asyncio.run(main())

