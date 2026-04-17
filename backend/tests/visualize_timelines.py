"""
Timeline Visualization — runs the full pipeline and exports results to Excel.

Usage:
    python tests/visualize_timelines.py

Outputs:
    tests/output/timelines.xlsx     — one sheet per field, entries as rows
    tests/output/summary.xlsx       — single sheet, all fields summarized
    tests/output/stats.xlsx         — pipeline stats + manual review + unconfirmed

Requires: OPENAI_API_KEY in .env
"""
from __future__ import annotations

import asyncio
import json
import os
import sys
from datetime import date

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()


def _style_header(ws, num_cols: int) -> None:
    """Apply header styling to first row."""
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        bottom=Side(style="thin", color="2F5496"),
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border


def _auto_width(ws) -> None:
    """Auto-fit column widths."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 60)


def export_timelines_to_excel(result: dict, output_dir: str) -> None:
    """Export pipeline result to Excel files."""
    os.makedirs(output_dir, exist_ok=True)

    # ── timelines.xlsx: one sheet per field ────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    for field_name in sorted(result["timelines"].keys()):
        entries = result["timelines"][field_name]
        ws = wb.create_sheet(title=field_name[:31])  # Excel 31 char limit

        headers = ["Date", "End Date", "Value", "Source", "Email ID", "Clause ID"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        _style_header(ws, len(headers))

        for row_idx, e in enumerate(entries, 2):
            ws.cell(row=row_idx, column=1, value=e.get("date"))
            ws.cell(row=row_idx, column=2, value=e.get("end_date", ""))
            ws.cell(row=row_idx, column=3, value=str(e.get("value", "")))
            ws.cell(row=row_idx, column=4, value=e.get("source", "")[:80])
            ws.cell(row=row_idx, column=5, value=e.get("email_source_id", ""))
            ws.cell(row=row_idx, column=6, value=e.get("clause_id", ""))

            # Alternate row shading
            if row_idx % 2 == 0:
                light_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0",
                                         fill_type="solid")
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = light_fill

        _auto_width(ws)

    timelines_path = os.path.join(output_dir, "timelines.xlsx")
    wb.save(timelines_path)
    print(f"Saved: {timelines_path} ({len(result['timelines'])} fields)")

    # ── summary.xlsx: all fields on one sheet ──────────────────────────
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "All Timelines"

    headers2 = ["Field", "Date", "End Date", "Value", "Source", "Email ID"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header(ws2, len(headers2))

    row_idx = 2
    for field_name in sorted(result["timelines"].keys()):
        for e in result["timelines"][field_name]:
            ws2.cell(row=row_idx, column=1, value=field_name)
            ws2.cell(row=row_idx, column=2, value=e.get("date"))
            ws2.cell(row=row_idx, column=3, value=e.get("end_date", ""))
            ws2.cell(row=row_idx, column=4, value=str(e.get("value", "")))
            ws2.cell(row=row_idx, column=5, value=e.get("source", "")[:80])
            ws2.cell(row=row_idx, column=6, value=e.get("email_source_id", ""))
            row_idx += 1

    _auto_width(ws2)
    summary_path = os.path.join(output_dir, "summary.xlsx")
    wb2.save(summary_path)
    print(f"Saved: {summary_path} ({row_idx - 2} total entries)")

    # ── stats.xlsx: pipeline metadata ──────────────────────────────────
    wb3 = Workbook()

    # Stats sheet
    ws_stats = wb3.active
    ws_stats.title = "Stats"
    stats = result["stats"]
    stat_headers = ["Metric", "Value"]
    for col, h in enumerate(stat_headers, 1):
        ws_stats.cell(row=1, column=col, value=h)
    _style_header(ws_stats, 2)
    for row_idx, (k, v) in enumerate(stats.items(), 2):
        ws_stats.cell(row=row_idx, column=1, value=k)
        ws_stats.cell(row=row_idx, column=2, value=v)
    _auto_width(ws_stats)

    # Manual review sheet
    ws_mr = wb3.create_sheet("Manual Review")
    mr_headers = ["Clause Text", "Reason", "Affected Field"]
    for col, h in enumerate(mr_headers, 1):
        ws_mr.cell(row=1, column=col, value=h)
    _style_header(ws_mr, len(mr_headers))
    for row_idx, item in enumerate(result["manual_review_items"], 2):
        ws_mr.cell(row=row_idx, column=1, value=item.get("clause_text", "")[:200])
        ws_mr.cell(row=row_idx, column=2, value=item.get("reason", ""))
        ws_mr.cell(row=row_idx, column=3, value=item.get("affected_field", ""))
    _auto_width(ws_mr)

    # Unconfirmed documents sheet
    ws_uc = wb3.create_sheet("Unconfirmed")
    uc_headers = ["Clause Text", "Email Source", "Doc Type"]
    for col, h in enumerate(uc_headers, 1):
        ws_uc.cell(row=1, column=col, value=h)
    _style_header(ws_uc, len(uc_headers))
    for row_idx, doc in enumerate(result["unconfirmed_documents"], 2):
        ws_uc.cell(row=row_idx, column=1, value=doc.get("clause_text", "")[:200])
        ws_uc.cell(row=row_idx, column=2, value=doc.get("email_source_id", ""))
        ws_uc.cell(row=row_idx, column=3, value=doc.get("doc_type", ""))
    _auto_width(ws_uc)

    # Assumptions sheet
    ws_as = wb3.create_sheet("Assumptions")
    ws_as.cell(row=1, column=1, value="Assumption")
    _style_header(ws_as, 1)
    for row_idx, a in enumerate(result["assumptions"], 2):
        ws_as.cell(row=row_idx, column=1, value=a)
    _auto_width(ws_as)

    stats_path = os.path.join(output_dir, "stats.xlsx")
    wb3.save(stats_path)
    print(f"Saved: {stats_path}")


async def main():
    from openai import AsyncOpenAI
    from engine.pipeline import start_session, evaluate, DB_PATH

    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        print("ERROR: OPENAI_API_KEY not set in .env or environment")
        sys.exit(1)

    client = AsyncOpenAI(api_key=api_key)

    print("Starting session...")
    session_id, extraction_results = start_session(DB_PATH)
    print(f"Session: {session_id}")
    print(f"Loaded {len(extraction_results)} extraction results")

    print("\nRunning full pipeline (evaluation_date=2026-06-01)...")
    result = await evaluate(
        session_id=session_id,
        full_email_dataset=[],
        evaluation_date_str="2026-06-01",
        openai_client=client,
    )

    print(f"\n--- Pipeline Stats ---")
    for k, v in result["stats"].items():
        print(f"  {k}: {v}")

    print(f"\n--- Timeline Fields ({len(result['timelines'])}) ---")
    for field in sorted(result["timelines"].keys()):
        entries = result["timelines"][field]
        print(f"  {field}: {len(entries)} entries")

    output_dir = os.path.join(os.path.dirname(__file__), "output")
    export_timelines_to_excel(result, output_dir)

    # Also save raw JSON for debugging
    json_path = os.path.join(output_dir, "result.json")
    with open(json_path, "w") as f:
        json.dump(result, f, indent=2, default=str)
    print(f"Saved: {json_path}")

    print("\nDone.")


if __name__ == "__main__":
    asyncio.run(main())
